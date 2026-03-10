from __future__ import annotations
import shutil
import tempfile
import html
import io
import json
import os
import re
import zipfile
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple
import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore

IGNORED_KEYS = {
    "library", "subContentId", "behaviour", "behavior", "copyright", "mime", "path",
    "width", "height", "expandImage", "minimizeImage", "contentName", "decorative",
    "useSeparator", "coverDescription", "defaultLanguage", "language", "embedTypes",
    "preloadedDependencies", "mainLibrary", "author", "source", "license", "yearsFrom",
    "yearsTo", "changes", "machineName", "majorVersion", "minorVersion", "patchVersion",
    "runnable", "coreApi", "fullscreen", "a11y", "disableImageZooming", "enableRetry",
    "showSolutionsRequiresInput", "showRetry", "showSolutions", "confirmCheckDialog",
    "confirmRetryDialog", "accessibility"
}

PRIORITY_TEXT_KEYS = {
    "text", "html", "description", "body", "summary", "question", "answer", "answers",
    "feedback", "solution", "instruction", "instructions", "caption", "title", "label",
    "header", "name", "statement", "intro", "prompt", "content", "tip", "placeholder"
}

NOISE_EXACT = {
    "Imagen", "Expandir Imagen", "Minimizar Imagen", "Sin título", "Untitled", "U",
    "true", "false", "null"
}

NOISE_PREFIXES = (
    "Sin título ", "Untitled ", "image/", "video/", "audio/",
)

EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")

class HTMLATexto(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.partes: List[str] = []

    def handle_starttag(self, tag: str, attrs: list) -> None:
        tag = (tag or "").lower()
        if tag == "br":
            self.partes.append("\n")
        elif tag in {"p", "div", "section", "article", "ul", "ol", "table", "blockquote"}:
            self.partes.append("\n")
        elif tag == "li":
            self.partes.append("\n- ")
        elif tag == "tr":
            self.partes.append("\n")
        elif tag in {"td", "th"}:
            if self.partes and not self.partes[-1].endswith(("\n", " | ")):
                self.partes.append(" | ")

    def handle_endtag(self, tag: str) -> None:
        tag = (tag or "").lower()
        if tag in {"p", "div", "section", "article", "blockquote"}:
            self.partes.append("\n")
        elif tag in {"li", "tr"}:
            self.partes.append("\n")

    def handle_data(self, data: str) -> None:
        if data:
            self.partes.append(data)

    def obtener_texto(self) -> str:
        texto = html.unescape("".join(self.partes)).replace("\xa0", " ")
        texto = texto.replace("\r\n", "\n").replace("\r", "\n")
        texto = re.sub(r"[ \t]+\n", "\n", texto)
        texto = re.sub(r"\n[ \t]+", "\n", texto)
        texto = re.sub(r"[ \t]{2,}", " ", texto)
        texto = re.sub(r"\n{3,}", "\n\n", texto)
        return texto.strip()


def html_a_texto(valor: str) -> str:
    parser = HTMLATexto()
    parser.feed(valor)
    parser.close()
    return parser.obtener_texto()


def decodificar_bytes(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def limpiar_texto_plano(texto: str) -> str:
    texto = html.unescape(texto).replace("\xa0", " ")
    texto = texto.replace("\r\n", "\n").replace("\r", "\n")
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


def normalize_for_dedup(texto: str) -> str:
    texto = texto.strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto.casefold()


def parece_ruta_o_binario(texto: str) -> bool:
    if re.fullmatch(r"[A-Za-z]:\\.*", texto):
        return True
    if re.fullmatch(
        r".*\.(png|jpg|jpeg|gif|svg|webp|bmp|mp4|avi|mov|mp3|wav|pdf|doc|docx|ppt|pptx|xls|xlsx)",
        texto,
        re.IGNORECASE,
    ):
        return True
    if "/" in texto and re.search(r"\.[A-Za-z0-9]{2,5}$", texto):
        return True
    return False


def es_ruido(texto: str) -> bool:
    texto = (texto or "").strip()
    if not texto:
        return True
    if texto in NOISE_EXACT:
        return True
    if any(texto.startswith(prefijo) for prefijo in NOISE_PREFIXES):
        return True
    if re.fullmatch(r"[A-Fa-f0-9-]{24,}", texto):
        return True
    if re.fullmatch(r"\d{1,4}", texto):
        return True
    if parece_ruta_o_binario(texto):
        return True
    return False


def sanitize_filename(name: str) -> str:
    name = name.strip().replace("\n", " ").replace("\r", " ")
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name or "archivo"


def unique_path(base_dir: Path, filename: str) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    candidate = base_dir / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    idx = 2
    while True:
        alt = base_dir / f"{stem}_{idx}{suffix}"
        if not alt.exists():
            return alt
        idx += 1


class RecolectorTexto:
    def __init__(self) -> None:
        self.bloques: List[str] = []
        self._vistos = set()

    def agregar(self, texto: str) -> None:
        if not texto:
            return
        texto = texto.strip()
        if not texto or es_ruido(texto):
            return
        clave = normalize_for_dedup(texto)
        if not clave or clave in self._vistos:
            return
        self._vistos.add(clave)
        self.bloques.append(texto)

    def recorrer(self, obj: Any) -> None:
        if isinstance(obj, dict):
            self._procesar_dict(obj)
        elif isinstance(obj, list):
            for item in obj:
                self.recorrer(item)
        elif isinstance(obj, str):
            texto = html_a_texto(obj) if ("<" in obj and ">" in obj) else limpiar_texto_plano(obj)
            self.agregar(texto)

    def _procesar_dict(self, obj: Dict[str, Any]) -> None:
        metadata = obj.get("metadata") if isinstance(obj.get("metadata"), dict) else None
        if metadata:
            content_type = str(metadata.get("contentType", "")).strip().lower()
            titulo = limpiar_texto_plano(str(metadata.get("title", "")).strip())
            if content_type == "page" and titulo and not es_ruido(titulo):
                self.agregar(titulo)

        for clave, valor in obj.items():
            clave_str = str(clave)
            clave_low = clave_str.lower()
            if clave_str in IGNORED_KEYS or clave_str == "metadata":
                continue
            if isinstance(valor, str) and clave_low in PRIORITY_TEXT_KEYS:
                texto = html_a_texto(valor) if ("<" in valor and ">" in valor) else limpiar_texto_plano(valor)
                self.agregar(texto)

        for clave, valor in obj.items():
            clave_str = str(clave)
            if clave_str in IGNORED_KEYS or clave_str == "metadata":
                continue
            if isinstance(valor, str):
                continue
            self.recorrer(valor)


def obtener_titulo_principal(zip_ref: zipfile.ZipFile) -> str:
    try:
        if "h5p.json" not in zip_ref.namelist():
            return ""
        data = json.loads(decodificar_bytes(zip_ref.read("h5p.json")))
        titulo = limpiar_texto_plano(str(data.get("title", "")).strip())
        return "" if es_ruido(titulo) else titulo
    except Exception:
        return ""


def iterar_archivos_contenido(nombres_zip: Iterable[str]) -> List[str]:
    candidatos = []
    for nombre in nombres_zip:
        nombre_low = nombre.lower()
        if not nombre_low.startswith("content/"):
            continue
        if nombre_low.endswith((".json", ".html", ".htm", ".txt")):
            candidatos.append(nombre)
    candidatos.sort(key=lambda x: (0 if x == "content/content.json" else 1, x.lower()))
    return candidatos


def extraer_texto_h5p_desde_bytes(data: bytes) -> Tuple[str, int]:
    with zipfile.ZipFile(io.BytesIO(data), "r") as zip_ref:
        recolector = RecolectorTexto()
        titulo_principal = obtener_titulo_principal(zip_ref)
        if titulo_principal:
            recolector.agregar(titulo_principal)
        candidatos = iterar_archivos_contenido(zip_ref.namelist())
        if not candidatos:
            raise ValueError("No se encontraron archivos de contenido dentro del paquete H5P.")
        for nombre in candidatos:
            contenido = decodificar_bytes(zip_ref.read(nombre))
            if nombre.lower().endswith(".json"):
                try:
                    data_obj = json.loads(contenido)
                    recolector.recorrer(data_obj)
                except json.JSONDecodeError:
                    texto = html_a_texto(contenido) if ("<" in contenido and ">" in contenido) else limpiar_texto_plano(contenido)
                    recolector.agregar(texto)
            else:
                texto = html_a_texto(contenido) if ("<" in contenido and ">" in contenido) else limpiar_texto_plano(contenido)
                recolector.agregar(texto)
        texto_final = "\n\n".join(recolector.bloques).strip()
        return texto_final, len(recolector.bloques)

def extraer_texto_h5p_desde_archivo(h5p_path: str) -> Tuple[str, int]:
    with zipfile.ZipFile(h5p_path, "r") as zip_ref:
        recolector = RecolectorTexto()
        titulo_principal = obtener_titulo_principal(zip_ref)
        if titulo_principal:
            recolector.agregar(titulo_principal)

        candidatos = iterar_archivos_contenido(zip_ref.namelist())
        if not candidatos:
            raise ValueError("No se encontraron archivos de contenido dentro del paquete H5P.")

        for nombre in candidatos:
            contenido = decodificar_bytes(zip_ref.read(nombre))
            if nombre.lower().endswith(".json"):
                try:
                    data_obj = json.loads(contenido)
                    recolector.recorrer(data_obj)
                except json.JSONDecodeError:
                    texto = html_a_texto(contenido) if ("<" in contenido and ">" in contenido) else limpiar_texto_plano(contenido)
                    recolector.agregar(texto)
            else:
                texto = html_a_texto(contenido) if ("<" in contenido and ">" in contenido) else limpiar_texto_plano(contenido)
                recolector.agregar(texto)

        texto_final = "\n\n".join(recolector.bloques).strip()
        return texto_final, len(recolector.bloques)

def _valor_celda_limpio(valor: Any) -> Any:
    if isinstance(valor, str):
        return valor.strip()
    return valor


def _fila_tiene_datos(fila: Tuple[Any, ...]) -> bool:
    return any(celda not in (None, "") for celda in fila)


def _normalizar_encabezado(valor: Any, indice: int) -> str:
    texto = "" if valor is None else str(valor).strip()
    return texto if texto else f"Columna_{indice}"


def _trim_fila_a_largo(fila: Tuple[Any, ...], largo: int) -> List[Any]:
    valores = list(fila[:largo])
    if len(valores) < largo:
        valores.extend([None] * (largo - len(valores)))
    return [_valor_celda_limpio(v) for v in valores]


def _extraer_tabla_desde_worksheet(ws) -> Tuple[List[str], List[List[Any]]]:
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return [], []
    indice_header = None
    for idx, fila in enumerate(filas):
        if _fila_tiene_datos(fila):
            indice_header = idx
            break
    if indice_header is None:
        return [], []
    fila_header = filas[indice_header]
    ultimo_no_vacio = 0
    for idx, val in enumerate(fila_header, start=1):
        if val not in (None, ""):
            ultimo_no_vacio = idx
    if ultimo_no_vacio == 0:
        return [], []
    headers = [_normalizar_encabezado(v, i + 1) for i, v in enumerate(fila_header[:ultimo_no_vacio])]
    header_normalizado = [str(h).strip().casefold() for h in headers]
    data_rows: List[List[Any]] = []
    for fila in filas[indice_header + 1:]:
        fila_recortada = _trim_fila_a_largo(fila, len(headers))
        if not any(v not in (None, "") for v in fila_recortada):
            continue
        fila_cmp = ["" if v is None else str(v).strip().casefold() for v in fila_recortada]
        if fila_cmp == header_normalizado:
            continue
        data_rows.append(fila_recortada)
    return headers, data_rows


def leer_reporte_excel_desde_bytes(data: bytes, nombre_origen: str = "") -> Tuple[List[str], List[List[Any]], str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl no está disponible para leer reportes H5P.")
    try:
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el Excel '{nombre_origen}': {exc}") from exc
    mejor_headers: List[str] = []
    mejor_rows: List[List[Any]] = []
    mejor_hoja = ""
    hojas = []
    if wb.sheetnames:
        if wb.active.title in wb.sheetnames:
            hojas.append(wb[wb.active.title])
        for nombre in wb.sheetnames:
            if nombre != wb.active.title:
                hojas.append(wb[nombre])
    for ws in hojas:
        headers, rows = _extraer_tabla_desde_worksheet(ws)
        if (len(rows), len(headers)) > (len(mejor_rows), len(mejor_headers)):
            mejor_headers, mejor_rows, mejor_hoja = headers, rows, ws.title
    wb.close()
    return mejor_headers, mejor_rows, mejor_hoja


def leer_reporte_excel_desde_archivo(ruta_excel: str, nombre_origen: str = "") -> Tuple[List[str], List[List[Any]], str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl no está disponible para leer reportes H5P.")
    try:
        wb = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el Excel '{nombre_origen or ruta_excel}': {exc}") from exc

    mejor_headers: List[str] = []
    mejor_rows: List[List[Any]] = []
    mejor_hoja = ""

    hojas = []
    if wb.sheetnames:
        if wb.active.title in wb.sheetnames:
            hojas.append(wb[wb.active.title])
        for nombre in wb.sheetnames:
            if nombre != wb.active.title:
                hojas.append(wb[nombre])

    for ws in hojas:
        headers, rows = _extraer_tabla_desde_worksheet(ws)
        if (len(rows), len(headers)) > (len(mejor_rows), len(mejor_headers)):
            mejor_headers, mejor_rows, mejor_hoja = headers, rows, ws.title

    wb.close()
    return mejor_headers, mejor_rows, mejor_hoja

def _listar_h5p_en_zip(nombres: Iterable[str]) -> List[str]:
    archivos = [n for n in nombres if not n.endswith("/") and n.lower().endswith(".h5p")]
    archivos.sort(key=lambda x: x.lower())
    return archivos


def _listar_reportes_excel_en_zip(nombres: Iterable[str]) -> List[str]:
    candidatos = []
    for nombre in nombres:
        nl = nombre.lower()
        if nombre.endswith("/"):
            continue
        if not nl.endswith(EXCEL_EXTENSIONS):
            continue
        candidatos.append(nombre)
    candidatos.sort(
        key=lambda x: (
            0 if "reporte_h5p" in x.lower() else 1,
            0 if os.path.basename(x).lower().startswith("reporte_h5p") else 1,
            x.lower(),
        )
    )
    return candidatos


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _stem_variants(filename: str) -> List[str]:
    filename = Path(filename).name
    stem = Path(filename).stem
    variants = {stem, filename}
    if stem.endswith(".h5p"):
        variants.add(Path(stem).stem)
    return [v.casefold() for v in variants if v]


def _content_id_from_name(name: str) -> str:
    stem = Path(name).stem
    m = re.match(r"^(\d{8,})[_-]", stem)
    return m.group(1) if m else ""


def _title_from_name(name: str) -> str:
    stem = Path(name).stem
    title = re.sub(r"^\d{8,}[_-]", "", stem).strip()
    return title or stem


def _build_report_dataframe(report_frames: List[pd.DataFrame]) -> pd.DataFrame:
    if not report_frames:
        return pd.DataFrame()
    all_columns: List[str] = []
    for df in report_frames:
        for col in df.columns.tolist():
            if col not in all_columns:
                all_columns.append(col)
    normalized = []
    for df in report_frames:
        tmp = df.copy()
        for col in all_columns:
            if col not in tmp.columns:
                tmp[col] = ""
        normalized.append(tmp[all_columns])
    return pd.concat(normalized, ignore_index=True, sort=False)


def _prepare_lookup(report_df: pd.DataFrame) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    by_local: Dict[str, Dict[str, Any]] = {}
    by_content: Dict[str, Dict[str, Any]] = {}
    if report_df.empty:
        return by_local, by_content
    for _, row in report_df.iterrows():
        row_map = {str(k): row[k] for k in report_df.columns}
        local_filename = _safe_str(row_map.get("local_filename"))
        if local_filename:
            for variant in _stem_variants(local_filename):
                by_local[variant] = row_map
        content_id = _safe_str(row_map.get("content_id"))
        if content_id:
            by_content[content_id.casefold()] = row_map
    return by_local, by_content


def _meta_from_report_row(report_row: Optional[Dict[str, Any]], fallback_name: str, zip_name: str) -> Dict[str, Any]:
    if report_row is None:
        content_id = _content_id_from_name(fallback_name)
        title = _title_from_name(fallback_name)
        return {
            "display_name": Path(fallback_name).name,
            "origin": "upload_zip_h5p",
            "source_url": "",
            "Archivo": content_id,
            "name": title,
            "link_class": "",
            "content_id": content_id,
            "title": title,
            "url": "",
            "zip_name": zip_name,
        }
    content_id = _safe_str(report_row.get("content_id")) or _content_id_from_name(fallback_name)
    title = _safe_str(report_row.get("title")) or _safe_str(report_row.get("name")) or _title_from_name(fallback_name)
    url = _safe_str(report_row.get("url"))
    return {
        "display_name": Path(fallback_name).name,
        "origin": "upload_zip_h5p",
        "source_url": "",
        "Archivo": content_id,
        "name": title,
        "link_class": url,
        "content_id": content_id,
        "title": title,
        "url": url,
        "zip_name": zip_name,
        "view_url": _safe_str(report_row.get("view_url")),
        "scope": _safe_str(report_row.get("scope")),
        "local_filename": _safe_str(report_row.get("local_filename")) or Path(fallback_name).name,
    }

def _write_unified_excel(report_df: pd.DataFrame, out_path: Path, warnings: Sequence[str], stats: Dict[str, Any]) -> Path:
    summary_rows = [
        {"Concepto": "Fecha de generación", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
        {"Concepto": "ZIP procesados", "Valor": stats.get("zip_count", 0)},
        {"Concepto": "H5P encontrados", "Valor": stats.get("h5p_count", 0)},
        {"Concepto": "TXT generados", "Valor": stats.get("txt_count", 0)},
        {"Concepto": "Reportes encontrados", "Valor": stats.get("report_count", 0)},
        {"Concepto": "Advertencias", "Valor": " | ".join(warnings) if warnings else "Sin advertencias"},
    ]
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if report_df.empty:
            pd.DataFrame(columns=["content_id", "title", "url"]).to_excel(writer, index=False, sheet_name="Reporte_Unificado")
        else:
            report_df.to_excel(writer, index=False, sheet_name="Reporte_Unificado")
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Resumen_Proceso")
    return out_path


def _bundle_outputs(txt_paths: Sequence[str], excel_path: Path, out_zip: Path) -> Path:
    with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for txt_path in txt_paths:
            p = Path(txt_path)
            if p.exists():
                zf.write(p, arcname=f"txt/{p.name}")
        if excel_path.exists():
            zf.write(excel_path, arcname=excel_path.name)
    return out_zip

def process_h5p_zip_uploads(
    zip_uploads: Sequence[Tuple[str, str]],
    work_dir: str,
) -> Dict[str, Any]:
    
    work_path = Path(work_dir)
    txt_dir = work_path / "h5p_txt"
    txt_dir.mkdir(parents=True, exist_ok=True)

    warnings: List[str] = []
    report_frames: List[pd.DataFrame] = []
    extracted_items: List[Dict[str, Any]] = []
    h5p_count = 0
    report_count = 0
    temp_extract_dir = work_path / "_tmp_h5p_extract"
    temp_extract_dir.mkdir(parents=True, exist_ok=True)

    for zip_name, zip_path in zip_uploads:
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                names = zf.namelist()
                h5p_items = _listar_h5p_en_zip(names)
                report_items = _listar_reportes_excel_en_zip(names)

                if not h5p_items:
                    warnings.append(f"No se encontraron H5P en {zip_name}")

                for report_item in report_items:
                    report_tmp = unique_path(
                        temp_extract_dir,
                        sanitize_filename(Path(report_item).name),
                    )
                    try:
                        with zf.open(report_item, "r") as src, open(report_tmp, "wb") as dst:
                            shutil.copyfileobj(src, dst, length=1024 * 1024)

                        headers, rows, _ = leer_reporte_excel_desde_archivo(
                            str(report_tmp),
                            f"{zip_name}::{report_item}",
                        )
                        if headers:
                            df = pd.DataFrame(rows, columns=headers)
                            df["__zip_name"] = zip_name
                            df["__report_name"] = Path(report_item).name
                            report_frames.append(df)
                            report_count += 1
                    except Exception as exc:
                        warnings.append(f"Error leyendo reporte {Path(report_item).name} en {zip_name}: {exc}")
                    finally:
                        try:
                            if report_tmp.exists():
                                report_tmp.unlink()
                        except Exception:
                            pass

                for h5p_item in h5p_items:
                    h5p_count += 1
                    h5p_tmp = unique_path(
                        temp_extract_dir,
                        sanitize_filename(Path(h5p_item).name),
                    )
                    try:
                        with zf.open(h5p_item, "r") as src, open(h5p_tmp, "wb") as dst:
                            shutil.copyfileobj(src, dst, length=1024 * 1024)

                        texto, bloques = extraer_texto_h5p_desde_archivo(str(h5p_tmp))
                        if not texto.strip():
                            warnings.append(f"Sin texto útil en {Path(h5p_item).name} ({zip_name})")
                            continue

                        txt_name = sanitize_filename(Path(h5p_item).stem + ".txt")
                        txt_path = unique_path(txt_dir, txt_name)
                        txt_path.write_text(texto + "\n", encoding="utf-8")

                        extracted_items.append(
                            {
                                "txt_path": str(txt_path),
                                "txt_name": txt_path.name,
                                "h5p_name": Path(h5p_item).name,
                                "zip_name": zip_name,
                                "text_blocks": bloques,
                            }
                        )
                    except Exception as exc:
                        warnings.append(f"Error procesando H5P {Path(h5p_item).name} en {zip_name}: {exc}")
                    finally:
                        try:
                            if h5p_tmp.exists():
                                h5p_tmp.unlink()
                        except Exception:
                            pass

        except zipfile.BadZipFile:
            warnings.append(f"ZIP inválido o corrupto: {zip_name}")
        except Exception as exc:
            warnings.append(f"Error procesando ZIP {zip_name}: {exc}")
            
    report_df = _build_report_dataframe(report_frames)
    by_local, by_content = _prepare_lookup(report_df)

    txt_meta: Dict[str, Dict[str, Any]] = {}
    txt_paths: List[str] = []
    report_links_rows: List[Dict[str, Any]] = []

    for item in extracted_items:
        txt_path = item["txt_path"]
        h5p_name = item["h5p_name"]
        report_row = None
        for variant in _stem_variants(h5p_name):
            report_row = by_local.get(variant)
            if report_row is not None:
                break
        if report_row is None:
            content_id = _content_id_from_name(h5p_name)
            if content_id:
                report_row = by_content.get(content_id.casefold())
        meta = _meta_from_report_row(report_row, h5p_name, item["zip_name"])
        txt_meta[txt_path] = meta
        txt_paths.append(txt_path)
        report_links_rows.append(
            {
                "content_id": meta.get("content_id", ""),
                "title": meta.get("title", ""),
                "url": meta.get("url", ""),
                "local_filename": meta.get("local_filename", Path(h5p_name).name),
                "txt_filename": Path(txt_path).name,
                "zip_name": item["zip_name"],
            }
        )

    report_links_df = pd.DataFrame(report_links_rows)
    if not report_df.empty:
        # de-duplicate by content_id prefiriendo el primer registro encontrado
        if "content_id" in report_df.columns:
            report_df = report_df.copy()
            report_df["content_id"] = report_df["content_id"].map(_safe_str)
            report_df = report_df.drop_duplicates(subset=["content_id", "local_filename"], keep="first")

    bundle_zip_path = ""
    unified_excel_path = ""
    if txt_paths or not report_df.empty:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        unified_excel = work_path / f"reporte_h5p_unificado_{ts}.xlsx"
        _write_unified_excel(
            report_df,
            unified_excel,
            warnings,
            {
                "zip_count": len(zip_uploads),
                "h5p_count": h5p_count,
                "txt_count": len(txt_paths),
                "report_count": report_count,
            },
        )
        unified_excel_path = str(unified_excel)
        bundle_zip = work_path / f"h5p_txt_unificado_{ts}.zip"
        _bundle_outputs(txt_paths, unified_excel, bundle_zip)
        bundle_zip_path = str(bundle_zip)

    return {
        "txt_paths": txt_paths,
        "txt_meta": txt_meta,
        "report_df": report_df,
        "report_links_df": report_links_df,
        "bundle_zip_path": bundle_zip_path,
        "unified_excel_path": unified_excel_path,
        "warnings": warnings,
        "stats": {
            "zip_count": len(zip_uploads),
            "h5p_count": h5p_count,
            "txt_count": len(txt_paths),
            "report_count": report_count,
        },
    }

def _fallback_extract_urls(text: str) -> List[str]:
    pattern = re.compile(r"(?:https?://|www\.)[^\s<>\"'\]\)]+", re.IGNORECASE)
    return [m.group(0).rstrip('.,;:!?') for m in pattern.finditer(text or "")]


def _extract_links_from_txt_h5p_bytes(
    txt_bytes: bytes,
    filename: str,
    *,
    meta: Optional[Dict[str, Any]] = None,
    url_extractor: Optional[Callable[[str], List[str]]] = None,
) -> List[Dict[str, Any]]:
    texto = decodificar_bytes(txt_bytes)
    extractor = url_extractor or _fallback_extract_urls
    rows: List[Dict[str, Any]] = []
    meta = meta or {}
    archivo = _safe_str(meta.get("Archivo") or meta.get("content_id"))
    nombre = _safe_str(meta.get("name") or meta.get("title"))
    link_class = _safe_str(meta.get("link_class") or meta.get("url"))
    source_url = _safe_str(meta.get("source_url"))

    for line in texto.splitlines():
        line_clean = line.strip()
        if not line_clean:
            continue
        urls = extractor(line_clean)
        if not urls:
            continue
        for url in urls:
            rows.append(
                {
                    "Nombre del Archivo": Path(filename).name,
                    "Archivo": archivo,
                    "name": nombre,
                    "link_class": link_class,
                    "source_url": source_url,
                    "Página/Diapositiva": "",
                    "Links": url,
                }
            )
    return rows

def run_h5p_txt_link_report_streamlit(
    txt_paths: Sequence[str],
    *,
    txt_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    progress_bar,
    status_text,
    url_extractor: Optional[Callable[[str], List[str]]] = None,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    if not txt_paths:
        empty = pd.DataFrame(
            columns=[
                "Nombre del Archivo",
                "Archivo",
                "name",
                "link_class",
                "source_url",
                "Página/Diapositiva",
                "Links",
            ]
        )
        return empty, []

    txt_meta = txt_meta or {}
    all_rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    total = len(txt_paths)
    for idx, path in enumerate(txt_paths, start=1):
        p = Path(path)
        status_text.markdown(f"Analizando H5P TXT **{idx}/{total}** · `{p.name}`")
        try:
            progress_bar.progress((idx - 1) / total if total else 0.0)
        except Exception:
            pass
        try:
            data = p.read_bytes()
            rows = _extract_links_from_txt_h5p_bytes(
                data,
                p.name,
                meta=txt_meta.get(path, {}),
                url_extractor=url_extractor,
            )
            all_rows.extend(rows)
        except Exception as exc:
            errors.append({"Archivo": p.name, "Error": str(exc)})
        try:
            progress_bar.progress(idx / total if total else 1.0)
        except Exception:
            pass

    if all_rows:
        df = pd.DataFrame(all_rows)
        df = df.sort_values(["Nombre del Archivo", "Página/Diapositiva", "Links"]).reset_index(drop=True)
    else:
        df = pd.DataFrame(
            columns=[
                "Nombre del Archivo",
                "Archivo",
                "name",
                "link_class",
                "source_url",
                "Página/Diapositiva",
                "Links",
            ]
        )
    return df, errors


